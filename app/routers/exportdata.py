import io
from datetime import datetime, timedelta
from typing import Any, Dict, List, Union

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import desc
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import ExportRequest, User
from app.utils import get_current_user_with_api_key

router = APIRouter(prefix="/export", tags=["Export Data"])


import io
from datetime import datetime, timedelta
from typing import Any, Dict, List, Union

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import ExportRequest, User
from app.utils import get_current_user_with_api_key

router = APIRouter(prefix="/export", tags=["Export Data"])


def create_comprehensive_excel(data: Dict[str, Dict[str, Any]]) -> io.BytesIO:
    """
    Generate a multi-sheet Excel file.
    
    Args:
        data: Structure:
        {
            "SheetName": {
                "type": "kv" | "table",
                "headers": ["Col1", "Col2"] (for table),
                "data": {...} (for kv) OR [{...}, ...] (for table)
            }
        }
    """
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=False, color="FFFFFF")
    
    for sheet_name, content in data.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        
        sheet_type = content.get("type")
        sheet_data = content.get("data")
        
        if sheet_type == "kv":
            # CASE 1: Key-Value Data (Profile)
            # No headers, just data
            
            # Data
            row_idx = 1
            if isinstance(sheet_data, dict):
                for key, value in sheet_data.items():
                    # Style the key column to look like a header/label
                    cell_key = ws.cell(row=row_idx, column=1, value=str(key))
                    cell_key.fill = header_fill
                    cell_key.font = header_font
                    
                    ws.cell(row=row_idx, column=2, value=str(value))
                    row_idx += 1
            
            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 50

        elif sheet_type == "table":
            # CASE 2: Tabular Data
            headers = content.get("headers", [])
            
            # Write Headers (even if data is empty)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Write Data
            if isinstance(sheet_data, list) and sheet_data:
                for row_idx, item in enumerate(sheet_data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        val = item.get(header)
                        # Clean Up
                        cell_value = str(val) if val is not None else ""
                        ws.cell(row=row_idx, column=col_idx, value=cell_value)
            else:
                # Optional: Write "No Records" if empty
                ws.cell(row=2, column=1, value="No records found")

            # Auto-adjust column widths
            for i, _ in enumerate(headers, 1):
                col_letter = get_column_letter(i)
                ws.column_dimensions[col_letter].width = 20

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


def collect_comprehensive_user_data(user: User, db: Session) -> Dict[str, Any]:
    """Collect all available user data with explicit headers."""
    
    # 1. User Profile
    profile_data = {
        "User ID": user.id,
        "Phone Number": user.phone_number,
        "Account Number": user.account_number,
        "Email": user.email or "N/A",
        "First Name": user.first_name or "N/A",
        "Last Name": user.last_name or "N/A",
        "Role": user.role.value,
        "Status": "Active" if user.is_active else "Inactive",
        "KYC Status": user.kyc_status.value,
        "KYC Document": user.kyc_document_type or "N/A",
        "Wallet Balance": f"{user.wallet_balance:,.2f}",
        "Loan Limit": f"{user.loan_limit:,.2f}",
        "Celo Address": user.celo_address or "N/A",
        "Sui Address": user.sui_address or "N/A",
        "Address": user.address or "N/A",
        "City": user.city or "N/A",
        "Country": user.country or "N/A",
        "Date of Birth": str(user.date_of_birth) if user.date_of_birth else "N/A",
        "Registration Date": user.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        "Last Update": user.updated_at.strftime("%Y-%m-%d %H:%M:%S") if user.updated_at else "N/A"
    }

    # 2. Transactions
    transactions = sorted(user.transactions, key=lambda x: x.created_at, reverse=True)
    txn_headers = [
        "ID", "Reference", "Type", "Status", "Amount", "Balance Before", 
        "Balance After", "Date", "Description", "Recipient ID", "Payment Link ID", "Loan ID"
    ]
    txn_data = []
    for t in transactions:
        txn_data.append({
            "ID": t.id,
            "Reference": t.reference,
            "Type": t.transaction_type.value,
            "Status": t.status.value,
            "Amount": f"{t.amount:,.2f}",
            "Balance Before": f"{t.balance_before:,.2f}",
            "Balance After": f"{t.balance_after:,.2f}",
            "Date": t.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "Description": t.description,
            "Recipient ID": t.recipient_id or "",
            "Payment Link ID": t.payment_link_id or "",
            "Loan ID": t.loan_id or ""
        })

    # 3. Loans
    loans = sorted(user.loans, key=lambda x: x.created_at, reverse=True)
    loan_headers = [
        "ID", "Status", "Principal", "Interest Rate", "Interest Amount", 
        "Total Repayment", "Amount Paid", "Approved At", "Due Date", "Created At"
    ]
    loan_data = []
    for l in loans:
        loan_data.append({
            "ID": l.id,
            "Status": l.status.value,
            "Principal": f"{l.principal_amount:,.2f}",
            "Interest Rate": f"{l.interest_rate}%",
            "Interest Amount": f"{l.interest_amount:,.2f}",
            "Total Repayment": f"{l.total_amount:,.2f}",
            "Amount Paid": f"{l.amount_paid:,.2f}",
            "Approved At": l.approved_at.strftime("%Y-%m-%d %H:%M:%S") if l.approved_at else "",
            "Due Date": l.due_date.strftime("%Y-%m-%d") if l.due_date else "",
            "Created At": l.created_at.strftime("%Y-%m-%d %H:%M:%S")
        })

    # 4. Investments
    investments = sorted(user.investments, key=lambda x: x.created_at, reverse=True)
    invest_headers = [
        "ID", "Amount", "Interest Rate (Annual)", "Period", 
        "Accumulated Interest", "Active", "Created At", "Last Update"
    ]
    invest_data = []
    for inv in investments:
        invest_data.append({
            "ID": inv.id,
            "Amount": f"{inv.amount:,.2f}",
            "Interest Rate (Annual)": f"{inv.interest_rate}%",
            "Period": inv.period.value,
            "Accumulated Interest": f"{inv.accumulated_interest:,.2f}",
            "Active": "Yes" if inv.is_active else "No",
            "Created At": inv.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "Last Update": inv.last_accrual_update.strftime("%Y-%m-%d %H:%M:%S")
        })

    # 5. Login Logs
    logs = sorted(user.login_logs, key=lambda x: x.created_at, reverse=True)
    log_headers = ["ID", "Method", "IP Address", "User Agent", "Time"]
    log_data = []
    for log in logs:
        log_data.append({
            "ID": log.id,
            "Method": log.login_method,
            "IP Address": log.ip_address,
            "User Agent": log.user_agent,
            "Time": log.created_at.strftime("%Y-%m-%d %H:%M:%S")
        })

    # 6. Payment Links
    links = sorted(user.payment_links, key=lambda x: x.created_at, reverse=True)
    link_headers = [
        "ID", "Code", "Status", "Amount", "Description", 
        "Paid By Used ID", "Expires At", "Created At"
    ]
    link_data = []
    for link in links:
        link_data.append({
            "ID": link.id,
            "Code": link.link_code,
            "Status": link.status.value,
            "Amount": f"{link.amount:,.2f}",
            "Description": link.description,
            "Paid By Used ID": link.paid_by_id or "",
            "Expires At": link.expires_at.strftime("%Y-%m-%d %H:%M:%S") if link.expires_at else "",
            "Created At": link.created_at.strftime("%Y-%m-%d %H:%M:%S")
        })
        
    # 7. OTP Records (Audit)
    otps = sorted(user.otp_records, key=lambda x: x.created_at, reverse=True)
    otp_headers = ["ID", "Phone", "Used", "Attempts", "Created At", "Expires At"]
    otp_data = []
    for otp in otps:
        otp_data.append({
            "ID": otp.id,
            "Phone": otp.phone_number,
            "Used": "Yes" if otp.is_used else "No",
            "Attempts": otp.attempts,
            "Created At": otp.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "Expires At": otp.expires_at.strftime("%Y-%m-%d %H:%M:%S")
        })

    return {
        "Profile": {
            "type": "kv",
            "data": profile_data
        },
        "Transactions": {
            "type": "table",
            "headers": txn_headers,
            "data": txn_data
        },
        "Loans": {
            "type": "table",
            "headers": loan_headers,
            "data": loan_data
        },
        "Investments": {
            "type": "table",
            "headers": invest_headers,
            "data": invest_data
        },
        "Payment Links": {
            "type": "table",
            "headers": link_headers,
            "data": link_data
        },
        "Login Logs": {
            "type": "table",
            "headers": log_headers,
            "data": log_data
        },
        "OTP History": {
            "type": "table",
            "headers": otp_headers,
            "data": otp_data
        }
    }


async def schedule_export_task(user_id: int, db: Session):
    """Background task to handle delayed export"""
    export_req = ExportRequest(
        user_id=user_id,
        scheduled_for=datetime.utcnow() + timedelta(hours=24),
        status="scheduled"
    )
    db.add(export_req)
    db.commit()


@router.get("/user/immediate")
def export_user_data_immediate(
    current_user: User = Depends(get_current_user_with_api_key),
    db: Session = Depends(get_db)
):
    """Export comprehensive user data immediately as a multi-sheet Excel file"""
    try:
        # Collect all data
        full_data = collect_comprehensive_user_data(current_user, db)
        
        # Generate Excel
        excel_buffer = create_comprehensive_excel(full_data)
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Report_{current_user.first_name}_{current_user.last_name}_{timestamp}.xlsx"
        
        # Return as streaming response
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export data: {str(e)}"
        )


@router.post("/user/schedule")
async def export_user_data_scheduled(
    background_tasks: BackgroundTasks,
    delay_hours: int = 24,
    current_user: User = Depends(get_current_user_with_api_key),
    db: Session = Depends(get_db)
):
    """Schedule user data export after specified delay"""
    if delay_hours < 0 or delay_hours > 168:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Delay must be between 0 and 168 hours"
        )
    
    try:
        background_tasks.add_task(schedule_export_task, current_user.id, db)
        scheduled_time = datetime.utcnow() + timedelta(hours=delay_hours)
        
        return {
            "message": "Export scheduled successfully",
            "user_id": current_user.id,
            "scheduled_for": scheduled_time.isoformat(),
            "status": "scheduled"
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to schedule export: {str(e)}"
        )


@router.get("/user/status")
def check_export_status(
    current_user: User = Depends(get_current_user_with_api_key),
    db: Session = Depends(get_db)
):
    """Check status of scheduled exports"""
    exports = db.query(ExportRequest).filter(
        ExportRequest.user_id == current_user.id
    ).order_by(ExportRequest.scheduled_for.desc()).all()
    
    return {
        "user_id": current_user.id,
        "exports": [
            {
                "id": exp.id,
                "scheduled_for": exp.scheduled_for.isoformat(),
                "status": exp.status,
                "created_at": exp.created_at.isoformat()
            }
            for exp in exports
        ]
    }